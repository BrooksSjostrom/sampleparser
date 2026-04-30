import pathlib
import struct, re
from datetime import datetime, timedelta
import os
import json
import csv
import argparse

SAMPLER_NAMES = {
    "203250017534": "Alvin",
    "203280017535": "Simon",
    "203280017536": "Theo"
}
EPOCH = datetime(2000, 1, 1)

def parse_file(filename):
    with open(os.path.expanduser(filename), "rb") as f:
        data = f.read()

    # Event type string table starts at offset 0x12
    event_types = []
    i = 0x12
    while i < len(data):
        end = data.find(b'\x00', i)
        if end == -1:
            break
        s = data[i:end].decode('latin-1', errors='replace')
        if not s or not all(c.isprintable() or c == ' ' for c in s):
            break
        event_types.append(s)
        i = end + 1

    # Records: 16 bytes each, sync marker 0xAAAA at start
    positions = [m.start() for m in re.finditer(re.escape(b'\xaa\xaa'), data)]
    records = []
    for idx, pos in enumerate(positions):
        rec = data[pos:pos + 16]
        ts = struct.unpack('<I', rec[2:6])[0]
        dt = EPOCH + timedelta(seconds=ts)
        seq = struct.unpack('>H', rec[6:8])[0]  # sample sequence number
        evt_idx = rec[8]
        b9, b10 = rec[9], rec[10]
        if evt_idx != 0:
            evt_name = event_types[evt_idx] if evt_idx < len(event_types) else f"UNKNOWN_{evt_idx}"
        elif b9 == 0x01 and b10 == 0x00:
            evt_name = 'Success'
        elif 0 < b10 < len(event_types):
            evt_name = event_types[b10]
        else:
            evt_name = 'Missed Sample'
        program = b9  # program number
        site = b10  # site number
        bottle = rec[11]  # bottle number
        volume = rec[12]  # sample volume in mL
        status = rec[13]  # status byte
        checksum = struct.unpack('<H', rec[14:16])[0]
        records.append((idx + 1, dt, seq, evt_idx, evt_name, program, site, bottle, volume, checksum))

    # ── 4. Save ──────────────────────────────────────────────────────
    jsonobj = []
    for (idx, dt, seq, evt_idx, evt_name, program, site, bottle, volume, checksum) in records:
        category, bg, fg = cat_style(evt_name)
        jsonobj.append({
            "number": idx,
            "date": dt.strftime(format="%Y-%m-%d"),
            "time": dt.strftime(format="%H:%M:%S"),
            "sample_number": seq,
            "result": evt_name,
            "category": category,
            "program": program,
            "bottle": bottle,
            "volume": volume,
            "checksum": checksum
        })
    return jsonobj


# ── 2. Category/color helper ─────────────────────────────────────
def cat_style(evt_name):
    if evt_name == 'Success':
        return 'Successful Sample', 'E2EFDA', '375623'
    elif evt_name == 'Bottle Full':
        return 'Bottle Status', 'DDEBF7', '1F4E79'
    elif evt_name != 'Other':
        return 'Missed Sample', 'FFC7CE', 'C00000'
    else:
        return 'Other', 'F2F2F2', '595959'


def create_spreadsheet(records):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # ── 3. Build workbook ────────────────────────────────────────────
    wb = Workbook()

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center')

    # ── Sheet 1: Sample Log ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Sample Log"

    headers = ['#', 'Date', 'Time', 'Sample #', 'Result', 'Category',
               'Program', 'Bottle', 'Volume (mL)', 'Checksum']
    col_widths = [5, 13, 9, 9, 18, 20,
                  9, 8, 12, 11]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for i, (num, dt, seq, evt_idx, evt_name,
            program, site, bottle, volume, checksum) in enumerate(records):
        row = i + 2
        category, bg, fg = cat_style(evt_name)
        fill = PatternFill('solid', start_color=bg)
        fnt = Font(name='Arial', size=10, color=fg)

        vals = [num, dt.date(), dt.time(), seq, evt_name,
                category, program, bottle, volume, checksum]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = fnt
            cell.border = border
            if col == 1: cell.alignment = Alignment(horizontal='center')
            if col == 2: cell.number_format = 'YYYY-MM-DD'
            if col == 3: cell.number_format = 'HH:MM:SS'

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(records) + 1}"

    # ── Sheet 2: Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'Sample Log Summary'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws2.merge_cells('A1:C1')

    info = [
        ('Total Samples', len(records)),
        ('Date Range (Start)', records[0][1].date()),
        ('Date Range (End)', records[-1][1].date()),
        ('Sampling Interval', '1 hour'),
        ('Program #', records[0][6]),
        ('Bottle #', records[0][7]),
        ('Volume per Sample (mL)', records[0][8]),
    ]
    for r, (label, val) in enumerate(info, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(name='Arial', bold=True, size=11)
        cell = ws2.cell(row=r, column=2, value=val)
        cell.font = Font(name='Arial', size=11)
        if hasattr(val, 'year'):  # date object
            cell.number_format = 'YYYY-MM-DD'

    from collections import Counter
    counts = Counter(e for _, _, _, _, e, *_ in records)

    ws2.cell(row=11, column=1, value='Result').font = header_font
    ws2.cell(row=11, column=1).fill = header_fill
    ws2.cell(row=11, column=2, value='Count').font = header_font
    ws2.cell(row=11, column=2).fill = header_fill
    ws2.cell(row=11, column=3, value='Category').font = header_font
    ws2.cell(row=11, column=3).fill = header_fill

    for r, (evt_name, count) in enumerate(counts.most_common(), 12):
        category, bg, fg = cat_style(evt_name)
        ws2.cell(row=r, column=1, value=evt_name).font = Font(name='Arial', size=10)
        ws2.cell(row=r, column=2, value=count).font = Font(name='Arial', size=10)
        ws2.cell(row=r, column=3, value=category).font = Font(name='Arial', size=10, color=fg)
        ws2.cell(row=r, column=3).fill = PatternFill('solid', start_color=bg)
        for col in range(1, 4):
            ws2.cell(row=r, column=col).border = border

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 20


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
                prog='python3 samplog.py',
                description='Lists the contents of a directory',
                epilog='Text at the bottom of help')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-d", "--directory")
    group.add_argument("-c", "--choose", action="store_true")
    parser.add_argument("-f", "--format", choices=["xlsx", "csv", "json", "print"], default="print")
    parser.add_argument("-e", "--exports", choices=["all", "recent"], default="recent")
    parser.add_argument("-o", "--output")

    args = parser.parse_args()

    # ── 1. Parse binary ──────────────────────────────────────────────
    if args.directory:
        data_directory = os.path.expanduser(args.directory)
    elif args.choose:
        from tkinter import Tk
        from tkinter.filedialog import askdirectory
        data_directory = askdirectory(title='Select \"Data\" folder')  # shows dialog box and return the path
    else:
        data_directory = os.path.expanduser("~/Documents/Duke/Other/Wastewater/sampledata/apr17")
    devices = [path for path in os.listdir(data_directory) if not path.startswith(".")]
    exports = {device: [] for device in devices}
    for device in devices:
        all_exports = [path for path in os.listdir(os.path.join(data_directory, device)) if not path.startswith(".")]
        if args.exports == "all":
            exports[device] = all_exports
        elif args.exports == "recent":
            exports[device] = [max(all_exports)]
    combined = {}
    for device in devices:
        # print(device)
        # print(exports[device])
        # export = str(max(sorted(exports[device], key=lambda x: int(x))))
        # print(export)
        combined[device] = {}
        for export in exports[device]:
            filename = os.path.join(data_directory, device, export, "SampLog.bin")
            combined[device][export] = parse_file(filename)

    if args.format == "json":
        outfile = args.output or "out.json"
        with open(outfile, "w") as f:
            json.dump(combined, f, indent=4)
            print("Dumped json")
    elif args.format == "xlsx":
        print("Not supported yet")
        raise NotImplementedError
        # wb = create_spreadsheet(records)
        # wb.save('SampLog.xlsx')
        # print("Saved SampLog.xlsx")
    elif args.format == "csv":
        outfile = args.output or "out.csv"
        with open(outfile, "w") as f:
            csvwriter = csv.writer(f)
            csvwriter.writerow(["Serial number", "Name", "Export"] + [str(x) for x in range(36)])
            for device in devices:
                for export in exports[device]:
                    row = [device, SAMPLER_NAMES[device], export] + ["1" if sample["result"] == "Success" else "0" for sample in combined[device][export]]
                    csvwriter.writerow(row)
    elif args.format == "print":
        for device in devices:
            print(f"Device: {device} ({SAMPLER_NAMES[device]})")
            for export in exports[device]:
                print(f"    Export: {export} ({EPOCH + timedelta(seconds=int(export))})")
                print("        ", "".join(["1" if sample["result"] == "Success" else "0" for sample in combined[device][export]]))
